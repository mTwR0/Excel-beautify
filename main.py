import openpyxl
from excel_functions import set_width
from excel_functions import change_headers_border
from excel_functions import change_text_border
from excel_functions import change_headers_text
from excel_functions import change_normal_text
from excel_functions import center_text
excel_file = r"your/excel/location.xlsx"
wb = openpyxl.load_workbook(excel_file)


column_width = 20

centrare_text="DA"

schimba_normal_text="DA"
normal_text_font = "Arial"
normal_text_color = "666666"
normal_text_fill = 'ffffff'
normal_text_size=10

normal_text_borders="DA"
normal_text_border="FULL" #EXTERIOR sau FULL
normal_text_border_style="thick" # dotted , dashed , double,  medium , thick , thin 
normal_text_border_color="0080ff"


schimba_headers_text="DA"
headers_text_font = "Calibri Light"
headers_text_color = "000000"
headers_text_fill = 'c2c2d6'
headers_text_size=14

headers_text_borders="DA"
headers_text_border="FULL" #EXTERIOR sau FULL
headers_border_style="thick" # dotted , dashed , double,  medium , thick , thin 
headers_border_color="4d001f"
#link pt borders : https://openpyxl.readthedocs.io/en/latest/_modules/openpyxl/styles/borders.html



for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    max_row = ws.max_row
    max_column = ws.max_column
    min_row=ws.min_row
    min_col=ws.min_column
    #pt headers --> font , color , fill , borders - full / outer

    set_width(ws, column_width)
    change_headers_border(ws,headers_text_borders,headers_text_border,headers_border_style,headers_border_color,min_row,min_col,max_column)
    change_text_border(normal_text_borders,normal_text_border,ws,min_row,max_row,min_col,max_column,normal_text_border_style,normal_text_border_color)
    change_headers_text(schimba_headers_text,headers_text_font,headers_text_size,headers_text_color,headers_text_fill,ws,min_row,min_col,max_column)
    change_normal_text(schimba_normal_text,normal_text_font,normal_text_size,normal_text_color,normal_text_fill,min_row,ws,max_row,min_col,max_column)
    center_text(centrare_text,ws,min_row,max_row,min_col,max_column)
    # Calculate the headers row (assuming headers are in the first row)

    #centreaza textul 


    wb.save(excel_file)
