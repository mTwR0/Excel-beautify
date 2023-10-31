#set width
import openpyxl
from openpyxl.styles import Border , Side
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Alignment


def set_width(sheet,desired_width):
    for column in sheet.columns:
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = desired_width

def change_headers_border(ws,headers_text_borders,headers_text_border,headers_border_style,headers_border_color,min_row,min_col,max_column):
    if headers_text_borders=="DA":
        if headers_text_border=="EXTERIOR":
            headers_cells = list(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_column))[0]    
            for cell in headers_cells:
                
                if cell.column==min_col:
                    top=Side(border_style=headers_border_style,color=headers_border_color)
                    bottom=Side(border_style=headers_border_style,color=headers_border_color)
                    left=Side(border_style=headers_border_style,color=headers_border_color)
                    #right=Side(border_style="dashed",color=headers_border_color)
                    border=Border(top=top,bottom=bottom,left=left)
                    cell.border=border
                elif cell.column==max_column:
                    top=Side(border_style=headers_border_style,color=headers_border_color)
                    bottom=Side(border_style=headers_border_style,color=headers_border_color)
                    #left=Side(border_style="dashed",color=headers_border_color)
                    right=Side(border_style=headers_border_style,color=headers_border_color)
                    border=Border(top=top,bottom=bottom,right=right)
                    cell.border=border
                else:
                    top=Side(border_style=headers_border_style,color=headers_border_color)
                    bottom=Side(border_style=headers_border_style,color=headers_border_color)
                    border=Border(top=top,bottom=bottom)
                    cell.border=border
                    #border left si right 
            #restul de borders 
            #pune borders
        elif headers_text_border == "FULL":
             headers_cells = list(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_column))[0]
             for cell in headers_cells:
                top=Side(border_style=headers_border_style,color=headers_border_color)
                bottom=Side(border_style=headers_border_style,color=headers_border_color)
                left=Side(border_style=headers_border_style,color=headers_border_color)
                right=Side(border_style=headers_border_style,color=headers_border_color)
                border=Border(top=top,bottom=bottom,left=left,right=right)
                cell.border=border
            #pune alt borders
    
def change_text_border(normal_text_borders,normal_text_border,ws,min_row,max_row,min_col,max_column,normal_text_border_style,normal_text_border_color):
    if normal_text_borders=="DA":
        if normal_text_border=="EXTERIOR":
             for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col, max_col=max_column):
                for cell in row:
                    if cell.row==max_row:
                        bottom=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                        left=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                        right=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                        border=Border(left=left,right=right,bottom=bottom)
                        cell.border=border
                    else:
                        left=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                        right=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                        border=Border(left=left,right=right)
                        cell.border=border
        elif normal_text_border == "FULL":
             for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col, max_col=max_column):
                for cell in row:
                    top=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                    bottom=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                    left=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                    right=Side(border_style=normal_text_border_style,color=normal_text_border_color)
                    border=Border(top=top,bottom=bottom,left=left,right=right)
                    cell.border=border
            #pune alt borders


def change_headers_text(schimba_headers_text,headers_text_font,headers_text_size,headers_text_color,headers_text_fill,ws,min_row,min_col,max_column):
    if schimba_headers_text == "DA":
        title_font = Font(name=headers_text_font, size=headers_text_size, color=headers_text_color,bold=True)
        title_fill = PatternFill(patternType='solid', fgColor=headers_text_fill)

        # Convert the generator to a list and access the first element (the headers row)
        headers_cells = list(ws.iter_rows(min_row=min_row, max_row=min_row, min_col=min_col, max_col=max_column))[0]
        
        # Apply styles to the headers row
        for cell in headers_cells:
            cell.font = title_font
            cell.fill = title_fill

def change_normal_text(schimba_normal_text,normal_text_font,normal_text_size,normal_text_color,normal_text_fill,min_row,ws,max_row,min_col,max_column):
        
    if schimba_normal_text =="DA":
        #pt contents --> font , color , borders - each cell / outer , fill 
        data_font = Font(name=normal_text_font, size=normal_text_size, color=normal_text_color)
        data_fill = PatternFill(patternType='solid', fgColor=normal_text_fill)

        # Apply styles to the data rows (A2:J<max_row>)
        for row in ws.iter_rows(min_row=min_row+1, max_row=max_row, min_col=min_col, max_col=max_column):
            for cell in row:
                cell.font = data_font
                cell.fill = data_fill

def center_text(centrare_text,ws,min_row,max_row,min_col,max_column):
    if centrare_text=="DA":
            
        center_alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_column):  # Adjust the range as needed
            for cell in row:
                cell.alignment = center_alignment
